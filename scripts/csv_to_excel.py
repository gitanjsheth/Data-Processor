import pandas as pd
import math
from openpyxl import Workbook

def reshape_column(column_data, num_columns=20):
    """Reshape a Series into rows with `num_columns` per row."""
    padded_len = math.ceil(len(column_data) / num_columns) * num_columns
    padded = column_data.tolist() + [''] * (padded_len - len(column_data))
    reshaped = [padded[i:i+num_columns] for i in range(0, len(padded), num_columns)]
    return reshaped

def csv_to_excel_multi_sheet(input_csv, output_excel):
    df = pd.read_csv(input_csv)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for col in df.columns:
        ws = wb.create_sheet(title=col[:31])  # Excel sheet name limit is 31 chars
        reshaped_data = reshape_column(df[col])

        for row_idx, row_data in enumerate(reshaped_data, start=1):
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(output_excel)
    print(f"Excel file saved as: {output_excel}")

# === USAGE ===
if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("Usage: python csv_to_multi_sheet_excel.py input.csv output.xlsx")
    else:
        csv_to_excel_multi_sheet(sys.argv[1], sys.argv[2])