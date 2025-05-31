"""Upload directory watcher that processes new files."""

import os
import time
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
import threading
from openpyxl import load_workbook

try:
    from .config import UPLOAD_DIR, WORKERS
    from .db import execute_query
    from .worker import FourBeltWorker
except ImportError:
    from config import UPLOAD_DIR, WORKERS
    from db import execute_query
    from worker import FourBeltWorker

class UploadHandler(FileSystemEventHandler):
    """Handler for file system events in the upload directory."""
    
    def __init__(self, executor):
        self.executor = executor
        self.processed_files = set()
        self.lock = threading.Lock()
    
    def on_created(self, event):
        """Handle file creation events."""
        if not event.is_directory:
            self._process_file(event.src_path)
    
    def on_moved(self, event):
        """Handle file move events."""
        if not event.is_directory:
            self._process_file(event.dest_path)
    
    def _process_file(self, file_path):
        """Process a new file."""
        file_path = Path(file_path)
        
        # Only process Excel files
        if file_path.suffix.lower() not in ['.xlsx', '.xls']:
            return
        
        # Avoid processing the same file multiple times
        with self.lock:
            if str(file_path) in self.processed_files:
                return
            self.processed_files.add(str(file_path))
        
        print(f"New file detected: {file_path}")
        
        # Wait a moment for file to be fully written
        time.sleep(2)
        
        # Submit for processing
        self.executor.submit(self._process_excel_file, file_path)
    
    def _process_excel_file(self, file_path: Path):
        """Process an Excel file by reading all sheets and spawning workers."""
        try:
            # Insert into source_files table
            file_id = self._register_source_file(file_path)
            
            # Get sheet names
            workbook = load_workbook(filename=file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            
            print(f"Processing {len(sheet_names)} sheets from {file_path.name}")
            
            # Process each sheet with a worker
            futures = []
            for sheet_name in sheet_names:
                future = self.executor.submit(self._process_sheet, file_path, sheet_name, file_id)
                futures.append(future)
            
            # Wait for all sheets to complete and collect results
            total_processed = 0
            total_errors = 0
            
            for future in futures:
                try:
                    result = future.result(timeout=300)  # 5 minute timeout per sheet
                    total_processed += result.get('processed_count', 0)
                    total_errors += result.get('error_count', 0)
                    print(f"Sheet '{result.get('sheet')}' completed: {result.get('processed_count')} records")
                except Exception as e:
                    print(f"Error processing sheet: {e}")
                    total_errors += 1
            
            # Update source file with results
            self._update_source_file_status(file_id, total_processed, total_errors)
            
            print(f"File {file_path.name} completed: {total_processed} records, {total_errors} errors")
            
        except Exception as e:
            print(f"Error processing file {file_path}: {e}")
            if 'file_id' in locals():
                self._update_source_file_status(file_id, 0, 1, str(e))
    
    def _process_sheet(self, file_path: Path, sheet_name: str, file_id: int):
        """Process a single sheet using FourBeltWorker."""
        worker = FourBeltWorker()
        return worker.run(str(file_path), sheet_name, file_id)
    
    def _register_source_file(self, file_path: Path) -> int:
        """Register a new source file in the database."""
        result = execute_query(
            """
            INSERT INTO source_files (filename, file_path, file_size, created_at, status)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id
            """,
            (
                file_path.name,
                str(file_path),
                file_path.stat().st_size,
                datetime.now(),
                'processing'
            )
        )
        return result[0][0]
    
    def _update_source_file_status(self, file_id: int, processed_count: int, error_count: int, error_message: str = None):
        """Update the status of a source file."""
        status = 'completed' if error_count == 0 else 'completed_with_errors'
        if processed_count == 0 and error_count > 0:
            status = 'failed'
        
        execute_query(
            """
            UPDATE source_files 
            SET status = %s, processed_count = %s, error_count = %s, 
                error_message = %s, processed_at = %s
            WHERE id = %s
            """,
            (status, processed_count, error_count, error_message, datetime.now(), file_id)
        )

class UploadWatcher:
    """Main class for watching the upload directory."""
    
    def __init__(self):
        self.observer = Observer()
        self.executor = ThreadPoolExecutor(max_workers=WORKERS)
        self.handler = UploadHandler(self.executor)
        
    def start(self):
        """Start watching the upload directory."""
        # Ensure upload directory exists
        Path(UPLOAD_DIR).mkdir(parents=True, exist_ok=True)
        
        print(f"Starting upload watcher on {UPLOAD_DIR} with {WORKERS} workers")
        
        # Process any existing files first
        self._process_existing_files()
        
        # Start watching for new files
        self.observer.schedule(self.handler, UPLOAD_DIR, recursive=False)
        self.observer.start()
        
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            self.stop()
    
    def stop(self):
        """Stop the watcher and cleanup."""
        print("Stopping upload watcher...")
        self.observer.stop()
        self.observer.join()
        self.executor.shutdown(wait=True)
        print("Upload watcher stopped")
    
    def _process_existing_files(self):
        """Process any Excel files that already exist in the upload directory."""
        upload_path = Path(UPLOAD_DIR)
        for file_path in upload_path.glob("*.xlsx"):
            if file_path.is_file():
                self.handler._process_file(file_path)
        
        for file_path in upload_path.glob("*.xls"):
            if file_path.is_file():
                self.handler._process_file(file_path)

def main():
    """Main entry point for the upload watcher."""
    watcher = UploadWatcher()
    watcher.start()

if __name__ == "__main__":
    main() 